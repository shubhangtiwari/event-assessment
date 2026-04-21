"""Flask app for the Claude Enablement pre-event assessment.

Routes:
    GET  /                       Survey form (or redirect to /thanks if already submitted)
    POST /submit                 Accept form submission, score, save, redirect to /thanks
    GET  /thanks                 Post-submission thank-you page
    GET  /admin/login            Admin password form
    POST /admin/login            Check password, set admin cookie
    GET  /admin/logout           Clear admin cookie
    GET  /admin/dashboard        Dashboard (requires admin cookie)
    GET  /admin/api/data         JSON: all submissions + current groups
    GET  /admin/api/export.xlsx  Download Excel workbook with groups
"""
from __future__ import annotations

import io
import secrets
from typing import Any

from flask import (
    Flask,
    Response,
    abort,
    jsonify,
    make_response,
    redirect,
    render_template,
    request,
    url_for,
)

from core import database as db
from core.config import (
    ADMIN_COOKIE,
    ADMIN_PASSWORD,
    ADMIN_PASSWORD_IS_DEFAULT,
    DEFAULT_GROUP_SIZE,
    GROUPING_SEED,
    HOST,
    PORT,
    RESPONDENT_COOKIE,
    SECRET_KEY_PATH,
)
from core.grouper import form_balanced_groups
from core.questions import (
    CORRECT_ANSWERS,
    REQUIRED_KEYS,
    SECTIONS,
)
from core.scorer import score_submission


# ---------------------------------------------------------------------------
# App setup
# ---------------------------------------------------------------------------

def _load_or_create_secret_key() -> bytes:
    """Persist a Flask SECRET_KEY across restarts so signed cookies survive."""
    if SECRET_KEY_PATH.exists():
        return SECRET_KEY_PATH.read_bytes()
    key = secrets.token_bytes(32)
    SECRET_KEY_PATH.write_bytes(key)
    try:
        SECRET_KEY_PATH.chmod(0o600)
    except OSError:
        pass
    return key


app = Flask(__name__)
app.config["SECRET_KEY"] = _load_or_create_secret_key()
app.config["SESSION_COOKIE_SAMESITE"] = "Lax"

db.init_db()


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _has_submitted(req) -> tuple[bool, str | None]:
    """Return (has_submitted, respondent_id)."""
    rid = req.cookies.get(RESPONDENT_COOKIE)
    if rid and db.respondent_exists(rid):
        return True, rid
    return False, None


def _is_admin(req) -> bool:
    token = req.cookies.get(ADMIN_COOKIE)
    return bool(token) and secrets.compare_digest(token, _admin_token())


def _admin_token() -> str:
    """Derive a stable admin token from SECRET_KEY + ADMIN_PASSWORD.

    Changes if either changes, so rotating the password invalidates old
    admin cookies without needing session storage.
    """
    import hashlib
    material = app.config["SECRET_KEY"] + ADMIN_PASSWORD.encode()
    return hashlib.sha256(material).hexdigest()


def _parse_submission(form) -> tuple[dict[str, Any], list[str]]:
    """Pull form fields into a normalized dict; return (data, errors)."""
    errors: list[str] = []
    data: dict[str, Any] = {}

    # Scalar fields.
    for key in ("name", "email", "role", "years_exp",
                "q5", "q6", "q8", "q9",
                "q11", "q12", "q13", "q14", "q15", "q16"):
        value = (form.get(key) or "").strip()
        data[key] = value

    # Likert -> int.
    for k in ("q5", "q6"):
        if data[k]:
            try:
                data[k] = int(data[k])
            except ValueError:
                errors.append(f"{k.upper()} must be 1–5.")
                data[k] = None
        else:
            data[k] = None

    # Multi-select.
    data["q7"] = form.getlist("q7")
    data["q10"] = form.getlist("q10")

    # Required fields.
    for key in REQUIRED_KEYS:
        v = data.get(key)
        if v in (None, ""):
            errors.append(f"{key} is required.")

    # Email sanity check.
    email = data.get("email", "")
    if email and ("@" not in email or "." not in email):
        errors.append("Please enter a valid email address.")

    return data, errors


def _build_groups(group_size: int | None = None) -> list[dict[str, Any]]:
    participants = db.fetch_participants()
    groups = form_balanced_groups(
        participants,
        group_size=group_size or DEFAULT_GROUP_SIZE,
        seed=GROUPING_SEED,
    )
    return [g.as_dict() for g in groups]


# ---------------------------------------------------------------------------
# Public routes (survey)
# ---------------------------------------------------------------------------

@app.route("/")
def survey():
    submitted, rid = _has_submitted(request)
    if submitted:
        return redirect(url_for("thanks"))
    return render_template("survey.html", sections=SECTIONS)


@app.route("/submit", methods=["POST"])
def submit():
    # Block duplicate submissions by cookie.
    submitted, _ = _has_submitted(request)
    if submitted:
        return redirect(url_for("thanks"))

    data, errors = _parse_submission(request.form)
    if errors:
        return render_template(
            "survey.html",
            sections=SECTIONS,
            errors=errors,
            submitted_values=data,
        ), 400

    # Also block by email (in case cookie was cleared).
    if db.email_exists(data["email"]):
        return render_template(
            "already_submitted.html",
            name=None,
            by_email=True,
        )

    # Score.
    score, level, breakdown = score_submission(data)

    # Persist.
    rid = db.insert_submission(
        name=data["name"],
        email=data["email"],
        role=data["role"],
        years_exp=data["years_exp"],
        answers=data,
        score=score,
        level=level,
        breakdown=breakdown,
    )

    # Re-compute groups so each submit genuinely triggers the grouper (the
    # result is also implicitly used on every dashboard view).
    _ = _build_groups()

    resp = make_response(redirect(url_for("thanks")))
    resp.set_cookie(
        RESPONDENT_COOKIE,
        rid,
        max_age=60 * 60 * 24 * 365,  # 1 year
        httponly=True,
        samesite="Lax",
    )
    return resp


@app.route("/thanks")
def thanks():
    submitted, rid = _has_submitted(request)
    if not submitted:
        return redirect(url_for("survey"))
    name = db.get_respondent_name(rid)
    return render_template("thanks.html", name=name)


# ---------------------------------------------------------------------------
# Admin routes
# ---------------------------------------------------------------------------

@app.route("/admin/login", methods=["GET", "POST"])
def admin_login():
    if request.method == "GET":
        return render_template("admin_login.html", default_password=ADMIN_PASSWORD_IS_DEFAULT)
    # POST
    supplied = (request.form.get("password") or "").strip()
    if not supplied or not secrets.compare_digest(supplied, ADMIN_PASSWORD):
        return render_template(
            "admin_login.html",
            error="Incorrect password.",
            default_password=ADMIN_PASSWORD_IS_DEFAULT,
        ), 401
    resp = make_response(redirect(url_for("admin_dashboard")))
    resp.set_cookie(
        ADMIN_COOKIE,
        _admin_token(),
        max_age=60 * 60 * 12,  # 12 hours
        httponly=True,
        samesite="Lax",
    )
    return resp


@app.route("/admin/logout")
def admin_logout():
    resp = make_response(redirect(url_for("admin_login")))
    resp.delete_cookie(ADMIN_COOKIE)
    return resp


@app.route("/admin/dashboard")
def admin_dashboard():
    if not _is_admin(request):
        return redirect(url_for("admin_login"))
    return render_template(
        "dashboard.html",
        default_group_size=DEFAULT_GROUP_SIZE,
    )


@app.route("/admin/api/data")
def admin_api_data():
    if not _is_admin(request):
        return jsonify(error="unauthorized"), 401

    group_size = request.args.get("group_size", default=DEFAULT_GROUP_SIZE, type=int)
    group_size = max(2, min(20, group_size))  # clamp

    responses = db.fetch_all_for_dashboard()
    groups = _build_groups(group_size=group_size)

    # Sanitize responses for the client: drop raw breakdown-json nesting,
    # keep the essentials the dashboard needs.
    trimmed = []
    for r in responses:
        trimmed.append({
            "respondent_id": r["respondent_id"],
            "submitted_at":  r["submitted_at"],
            "name":          r["name"],
            "email":         r["email"],
            "role":          r["role"],
            "years_exp":     r["years_exp"],
            "score":         r["score"],
            "level":         r["level"],
            "breakdown":     r["breakdown"],
        })

    level_counts: dict[str, int] = {}
    for r in trimmed:
        level_counts[r["level"]] = level_counts.get(r["level"], 0) + 1

    return jsonify({
        "responses":   trimmed,
        "groups":      groups,
        "total":       len(trimmed),
        "group_size":  group_size,
        "level_counts": level_counts,
    })


@app.route("/admin/api/settings", methods=["GET", "POST"])
def admin_api_settings():
    if not _is_admin(request):
        return jsonify(error="unauthorized"), 401
    if request.method == "GET":
        return jsonify(proxy_url=db.get_setting("proxy_url", ""))
    payload = request.get_json(silent=True) or {}
    proxy_url = (payload.get("proxy_url") or "").strip()
    db.set_setting("proxy_url", proxy_url)
    return jsonify(ok=True, proxy_url=proxy_url)


@app.route("/banner")
def banner():
    import segno
    proxy_url = db.get_setting("proxy_url", "").strip()
    qr_svg = ""
    if proxy_url:
        qr = segno.make(proxy_url, error="h")
        buf = io.BytesIO()
        qr.save(
            buf, kind="svg", scale=1, border=2, xmldecl=False, svgns=False,
            svgclass="qr-svg", lineclass=None, omitsize=True,
        )
        svg_str = buf.getvalue().decode("utf-8")
        width, height = qr.symbol_size(scale=1, border=2)
        qr_svg = svg_str.replace(
            "<svg ",
            f'<svg viewBox="0 0 {width} {height}" preserveAspectRatio="xMidYMid meet" ',
            1,
        )
    return render_template("banner.html", proxy_url=proxy_url, qr_svg=qr_svg)


@app.route("/admin/api/delete/<respondent_id>", methods=["POST", "DELETE"])
def admin_api_delete(respondent_id: str):
    if not _is_admin(request):
        return jsonify(error="unauthorized"), 401
    deleted = db.delete_submission(respondent_id)
    if not deleted:
        return jsonify(error="not_found"), 404
    return jsonify(ok=True, respondent_id=respondent_id)


@app.route("/admin/api/export.xlsx")
def admin_api_export():
    if not _is_admin(request):
        abort(401)

    # Lazy import so the rest of the app doesn't need openpyxl to run.
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    group_size = request.args.get("group_size", default=DEFAULT_GROUP_SIZE, type=int)
    responses = db.fetch_all_for_dashboard()
    groups = _build_groups(group_size=group_size)

    wb = Workbook()

    HEADER_FILL = PatternFill("solid", start_color="E20074")
    HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    LEVEL_FILL = {
        "Beginner":     PatternFill("solid", start_color="FFE4E1"),
        "Intermediate": PatternFill("solid", start_color="FFF4CE"),
        "Advanced":     PatternFill("solid", start_color="D9EAD3"),
    }

    def style_header(ws, row=1):
        for cell in ws[row]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")

    def autosize(ws):
        for col in ws.columns:
            letter = get_column_letter(col[0].column)
            w = max((len(str(c.value)) for c in col if c.value is not None), default=10)
            ws.column_dimensions[letter].width = min(max(w + 2, 10), 40)

    # Participants sheet.
    ws1 = wb.active
    ws1.title = "Participants"
    ws1.append(["Name", "Email", "Role", "Experience", "Score", "Level", "Group", "Submitted at"])
    style_header(ws1)
    ws1.freeze_panes = "A2"
    rid_to_group = {m["respondent_id"]: g["number"] for g in groups for m in g["members"]}
    for r in sorted(responses, key=lambda r: rid_to_group.get(r["respondent_id"], 999)):
        ws1.append([
            r["name"], r["email"], r["role"] or "", r["years_exp"] or "",
            r["score"], r["level"],
            rid_to_group.get(r["respondent_id"], ""),
            r["submitted_at"],
        ])
    # Level color-coding.
    for row_idx in range(2, ws1.max_row + 1):
        cell = ws1.cell(row=row_idx, column=6)
        if cell.value in LEVEL_FILL:
            cell.fill = LEVEL_FILL[cell.value]
    autosize(ws1)

    # Group assignments.
    ws2 = wb.create_sheet("Group Assignments")
    ws2.append(["Group", "Seat", "Name", "Email", "Role", "Level", "Score"])
    style_header(ws2)
    ws2.freeze_panes = "A2"
    for g in groups:
        for seat, m in enumerate(g["members"], start=1):
            ws2.append([g["number"], seat, m["name"], m["email"], m["role"], m["level"], m["score"]])
        ws2.append([""] * 7)
    for row in ws2.iter_rows(min_row=2):
        cell = row[5]
        if cell.value in LEVEL_FILL:
            cell.fill = LEVEL_FILL[cell.value]
    autosize(ws2)

    # Group stats.
    ws3 = wb.create_sheet("Group Stats")
    ws3.append(["Group", "Size", "Avg score", "Beginners", "Intermediate", "Advanced"])
    style_header(ws3)
    ws3.freeze_panes = "A2"
    for g in groups:
        mix = g["level_mix"]
        ws3.append([
            g["number"], g["size"], g["avg_score"],
            mix.get("Beginner", 0), mix.get("Intermediate", 0), mix.get("Advanced", 0),
        ])
    autosize(ws3)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        buf.read(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="groups.xlsx"'},
    )


# ---------------------------------------------------------------------------
# Entrypoint
# ---------------------------------------------------------------------------

def _startup_banner() -> None:
    border = "─" * 68
    print()
    print(f"┌{border}┐")
    print(f"│  Claude Enablement — pre-event assessment{' ' * 25}│")
    print(f"│  Bonn · April 20–21, 2026{' ' * 41}│")
    print(f"├{border}┤")
    print(f"│  Survey:    http://localhost:{PORT}/{' ' * 35}│")
    print(f"│  Dashboard: http://localhost:{PORT}/admin/login{' ' * 24}│")
    print(f"├{border}┤")
    if ADMIN_PASSWORD_IS_DEFAULT:
        print("│  ⚠  Admin password is the default ('admin').                       │")
        print("│     Before exposing publicly, set: export ADMIN_PASSWORD=...       │")
    else:
        print("│  Admin password: configured via ADMIN_PASSWORD env var.            │")
    print(f"│  Submissions so far: {db.count_submissions()}{' ' * (46 - len(str(db.count_submissions())))}│")
    print(f"└{border}┘")
    print()


if __name__ == "__main__":
    _startup_banner()
    # debug=False so the auto-reloader doesn't double the banner or the DB init.
    app.run(host=HOST, port=PORT, debug=False)
